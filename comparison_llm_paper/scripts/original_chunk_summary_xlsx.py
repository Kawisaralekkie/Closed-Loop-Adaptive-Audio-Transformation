#!/usr/bin/env python3
"""Turn original_chunk_summary.csv into an .xlsx workbook with 3 sheets.

Sheets:
    all      — every chunk
    speech   — chunks WITH speech      (had_speech=True)
    env       — chunks with NO speech   (had_speech=False)

had_speech is looked up from run_metrics_per_chunk.csv, keyed by
(source_file, chunk_index) for a chosen mode (default: fixed — the original
audio is identical across modes, and fixed never alters routing).

A `had_speech` column is added to every row. Header row is bold + frozen.

Usage:
    python3 scripts/original_chunk_summary_xlsx.py \
        logs/s3/20260701_171530/original_chunk_summary.csv \
        logs/s3/20260701_171530/run_metrics_per_chunk.csv \
        --out logs/s3/20260701_171530/original_chunk_summary.xlsx
"""

from __future__ import annotations

import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

NUMERIC_HINTS = ("amp_", "spec_", "freq_", "n_samples", "n_freq_bins")


def _build_speech_lookup(per_chunk_csv: str, mode_filter: str) -> dict:
    lut = {}
    with open(per_chunk_csv, newline="") as f:
        for r in csv.DictReader(f):
            if r.get("mode", "") != mode_filter:
                continue
            key = (r.get("source_file", ""), str(r.get("chunk_index", "")))
            lut[key] = str(r.get("had_speech", "")).strip().lower() in ("true", "1")
    return lut


def _coerce(header: str, value: str):
    """Convert numeric-looking cells to float/int so Excel treats them as numbers."""
    if value is None or value == "":
        return value
    if any(header.startswith(h) or header == h for h in NUMERIC_HINTS):
        try:
            f = float(value)
            return int(f) if header in ("n_samples", "n_freq_bins") else f
        except ValueError:
            return value
    return value


def _write_sheet(ws, header: list[str], rows: list[dict]) -> None:
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append([_coerce(h, r.get(h, "")) for h in header])
    # Auto-ish column widths.
    for i, h in enumerate(header, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)


def main() -> None:
    ap = argparse.ArgumentParser(description="original_chunk_summary.csv -> xlsx (3 sheets)")
    ap.add_argument("summary_csv", help="original_chunk_summary.csv")
    ap.add_argument("per_chunk_csv", help="run_metrics_per_chunk.csv (for had_speech)")
    ap.add_argument("--mode", default="fixed", help="Mode used for had_speech lookup (default: fixed)")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    speech_lut = _build_speech_lookup(args.per_chunk_csv, args.mode)

    with open(args.summary_csv, newline="") as f:
        reader = csv.DictReader(f)
        base_header = list(reader.fieldnames or [])
        rows = list(reader)

    header = base_header + ["had_speech"]
    all_rows, speech_rows, env_rows = [], [], []
    unmatched = 0
    for r in rows:
        key = (r.get("source_file", ""), str(r.get("chunk_index", "")))
        has = speech_lut.get(key)
        r = dict(r)
        r["had_speech"] = "" if has is None else ("True" if has else "False")
        all_rows.append(r)
        if has is True:
            speech_rows.append(r)
        elif has is False:
            env_rows.append(r)
        else:
            unmatched += 1

    out_path = os.path.abspath(args.out) if args.out \
        else os.path.splitext(os.path.abspath(args.summary_csv))[0] + ".xlsx"

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "all"
    _write_sheet(ws_all, header, all_rows)
    _write_sheet(wb.create_sheet("speech"), header, speech_rows)
    _write_sheet(wb.create_sheet("env"), header, env_rows)
    wb.save(out_path)

    print(f"all={len(all_rows)}  speech={len(speech_rows)}  env={len(env_rows)}"
          f"  unmatched={unmatched}")
    print(f"  \u2713 {out_path}")


if __name__ == "__main__":
    main()
