#!/usr/bin/env python3
"""Per-chunk privacy / preserve / psychoacoustic metrics -> xlsx (3 sheets).

Reads run_metrics_per_chunk.csv (11,130 rows = 2226 chunks x 5 modes) and
writes an Excel workbook with three sheets:

    overall — every chunk (all 5 modes)
    speech  — chunks WITH speech      (had_speech=True)
    env     — chunks with NO speech   (had_speech=False)

Every sheet keeps the mode column so all 5 modes are present side by side.

Columns kept (identifiers + requested metrics):
    mode, source_file, chunk_index, had_speech, speech_ratio,
    wer, cer, speaker_privacy, content_privacy, privacy_score,
    preserve_score, s_loud, s_hf, s_sc, s_con, s_psy,
    short_term_loudness, sharpness_proxy, roughness_proxy, fluctuation_proxy

Usage:
    python3 scripts/metrics_by_speech_xlsx.py \
        logs/s3/20260701_171530/run_metrics_per_chunk.csv \
        --out logs/s3/20260701_171530/metrics_by_speech.xlsx
"""

from __future__ import annotations

import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ID_COLS = ["mode", "source_file", "chunk_index", "had_speech", "speech_ratio"]
METRIC_COLS = [
    "wer", "cer", "speaker_privacy", "content_privacy", "privacy_score",
    "preserve_score", "s_loud", "s_hf", "s_sc", "s_con", "s_psy",
    "short_term_loudness", "sharpness_proxy", "roughness_proxy", "fluctuation_proxy",
]
NUMERIC = set(METRIC_COLS) | {"speech_ratio", "chunk_index"}


def _coerce(col: str, v: str):
    if v is None or v == "":
        return v
    if col in NUMERIC:
        try:
            fv = float(v)
            return int(fv) if col == "chunk_index" else fv
        except ValueError:
            return v
    return v


def _write_sheet(ws, header, rows):
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append([_coerce(h, r.get(h, "")) for h in header])
    for i, h in enumerate(header, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)


def main() -> None:
    ap = argparse.ArgumentParser(description="Per-chunk metrics -> xlsx (overall/speech/env)")
    ap.add_argument("per_chunk_csv")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    with open(args.per_chunk_csv, newline="") as f:
        reader = csv.DictReader(f)
        available = reader.fieldnames or []
        rows = list(reader)

    header = [c for c in (ID_COLS + METRIC_COLS) if c in available]
    missing = [c for c in (ID_COLS + METRIC_COLS) if c not in available]
    if missing:
        print(f"  note: columns not found and skipped: {missing}")

    def _is_true(r):
        return str(r.get("had_speech", "")).strip().lower() in ("true", "1")

    speech = [r for r in rows if _is_true(r)]
    env = [r for r in rows if not _is_true(r)]

    out_path = os.path.abspath(args.out) if args.out \
        else os.path.splitext(os.path.abspath(args.per_chunk_csv))[0] + "_metrics_by_speech.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "overall"
    _write_sheet(ws, header, rows)
    _write_sheet(wb.create_sheet("speech"), header, speech)
    _write_sheet(wb.create_sheet("env"), header, env)
    wb.save(out_path)

    print(f"overall={len(rows)}  speech={len(speech)}  env={len(env)}  cols={len(header)}")
    print(f"  \u2713 {out_path}")


if __name__ == "__main__":
    main()
